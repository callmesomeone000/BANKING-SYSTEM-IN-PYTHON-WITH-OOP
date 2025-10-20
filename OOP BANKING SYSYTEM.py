import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os

# ======================== Excel Setup ==========================
excel_file = "bank_records.xlsx"

if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "BankData"
    ws.append(["Name", "Age", "Gender", "Action", "Amount", "Balance"])
    wb.save(excel_file)

# ======================== Classes ===============================
class User:
    def __init__(self, name, age, gender):
        self.name = name
        self.age = age
        self.gender = gender

    def show_details(self):
        return f"Name: {self.name}\nAge: {self.age}\nGender: {self.gender}"


class Bank(User):
    def __init__(self, name, age, gender):
        super().__init__(name, age, gender)
        self.balance = 0

    def deposit(self, amount):
        self.balance += amount
        self.save_to_excel("Deposit", amount)
        return f"✅ Deposited ${amount}. Current Balance: ${self.balance}"

    def withdraw(self, amount):
        if amount > self.balance:
            return f"❌ Insufficient funds! Available: ${self.balance}"
        self.balance -= amount
        self.save_to_excel("Withdraw", amount)
        return f"💸 Withdrawn ${amount}. Remaining Balance: ${self.balance}"

    def save_to_excel(self, action, amount):
        wb = load_workbook(excel_file)
        ws = wb.active
        ws.append([self.name, self.age, self.gender, action, amount, self.balance])
        wb.save(excel_file)


# ======================== GUI Setup ==============================
root = tk.Tk()
root.title("🏦 Simple Bank System")
root.geometry("400x500")
root.config(bg="#1e1e1e")

# Styling
label_style = {"bg": "#1e1e1e", "fg": "white", "font": ("Arial", 11)}
entry_style = {"font": ("Arial", 11), "width": 20, "bg": "#f5f5f5"}
btn_style = {"font": ("Arial", 11, "bold"), "bg": "#2e8b57", "fg": "white", "width": 15, "relief": "raised"}

# Variables
name_var = tk.StringVar()
age_var = tk.StringVar()
gender_var = tk.StringVar()
amount_var = tk.StringVar()
user_obj = None

# ======================== Functions ==============================
def create_user():
    global user_obj
    name = name_var.get()
    age = age_var.get()
    gender = gender_var.get()

    if not (name and age and gender):
        messagebox.showwarning("Input Error", "Please fill all details.")
        return

    user_obj = Bank(name, age, gender)
    messagebox.showinfo("Success", f"👤 Account created for {name}!")
    clear_entries()


def deposit_money():
    if not user_obj:
        messagebox.showwarning("Error", "Create an account first.")
        return

    try:
        amount = float(amount_var.get())
        msg = user_obj.deposit(amount)
        messagebox.showinfo("Deposit", msg)
        clear_amount()
    except ValueError:
        messagebox.showerror("Invalid Input", "Enter a valid amount.")


def withdraw_money():
    if not user_obj:
        messagebox.showwarning("Error", "Create an account first.")
        return

    try:
        amount = float(amount_var.get())
        msg = user_obj.withdraw(amount)
        messagebox.showinfo("Withdraw", msg)
        clear_amount()
    except ValueError:
        messagebox.showerror("Invalid Input", "Enter a valid amount.")


def check_balance():
    if not user_obj:
        messagebox.showwarning("Error", "Create an account first.")
        return
    messagebox.showinfo("Balance", f"💰 Current Balance: ${user_obj.balance}")


def clear_entries():
    name_var.set("")
    age_var.set("")
    gender_var.set("")


def clear_amount():
    amount_var.set("")


# ======================== GUI Layout ==============================
tk.Label(root, text="🏦 Bank Account Form", bg="#1e1e1e", fg="lightgreen", font=("Arial", 16, "bold")).pack(pady=10)

tk.Label(root, text="Name:", **label_style).pack()
tk.Entry(root, textvariable=name_var, **entry_style).pack(pady=5)

tk.Label(root, text="Age:", **label_style).pack()
tk.Entry(root, textvariable=age_var, **entry_style).pack(pady=5)

tk.Label(root, text="Gender:", **label_style).pack()
tk.Entry(root, textvariable=gender_var, **entry_style).pack(pady=5)

tk.Button(root, text="Create Account", command=create_user, **btn_style).pack(pady=10)

tk.Label(root, text="Enter Amount:", **label_style).pack(pady=5)
tk.Entry(root, textvariable=amount_var, **entry_style).pack(pady=5)

tk.Button(root, text="Deposit", command=deposit_money, **btn_style).pack(pady=5)
tk.Button(root, text="Withdraw", command=withdraw_money, **btn_style).pack(pady=5)
tk.Button(root, text="Check Balance", command=check_balance, **btn_style).pack(pady=10)

tk.Label(root, text="All records are saved to bank_records.xlsx", bg="#1e1e1e", fg="gray", font=("Arial", 9)).pack(side="bottom", pady=10)

root.mainloop()
